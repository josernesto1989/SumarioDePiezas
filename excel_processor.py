import openpyxl

class ExcelProcessor:
    """
    Clase para procesar archivos de Excel (versiones .xlsx).
    """
    def __init__(self):
        # Puedes inicializar configuraciones o servicios aquí.
        pass

    def process_file(self, file_path_list, output_directory):
        tipos_de_piezas =["ASS","LCD", "LCD GALAXY TAB", "LCD NO ANDROID", 'LCD TAB 7"', 'LCD TAB 9"', 'LCD TAB 10"', 'TOUCH AND', 'TOUCH GTAB', 'TOUCH IPAD', 'TOUCH NO ANDROID', 'TOUCH TAB 8"', 'TOUCH TAB 7.85"', 'TOUCH TAB 7"', 'TOUCH TAB 9"', 'TOUCH TAB 10"', 'ANILLOS PERS', 'BAT DET', 'BAT', 'BAT INTERNA', 'CD', 'CARGADOR PF', 'CARGADOR UNIV', 'COV PERS', 'COV', 'COV DET', 'GEVEY', 'ML BT', 'ML CABLE', 'MEM USB 16GB', 'MCP AAA', 'MCP FIBRA CARBON', 'MCP LARGE', 'MCP MEDIUM', 'MCP R', 'MCP XLARGE', 'MCP XSMALL', 'MC', 'MSD 8GB', 'MSD 16GB', 'MSD 32GB', 'MSD 64GB', 'OTROS', 'PTA 1A', 'PTA 2A', 'RELOJ', 'TARJETA INT', 'TELEF', 'CAMARAS', 'BOCINAS', 'BOTONES, JOY, BSIM', 'CRISTALES', 'FLEX POWER, VOL, ETC', 'FLEX', 'PC CONECTOR 1', 'PC CONECTOR 2', 'PC CONECTOR 3', 'PC CONECTOR 4', 'PSIM', 'TAPAS TRASERAS', 'PC FLEX', 'PLACAS', 'BAT ADAP', 'ANILLOS UNIV', 'TOUCH REC', 'LCD REC', 'ASS SIN CRISTAL', 'MC SIN CLASIF', 'CD MAGNETICO', 'MSD 128GB', 'MSD 256GB', 'MEM USB 32GB', 'MEM USB 64GB', 'MEM USB 128GB', 'COV SIN CLASIF', 'MICROFONOS', 'COV LENTO MOV', 'CRISTALES CAM', 'BAT COMO ADAP', 'BAT DET COMO ADAP', 'MHG']
        total_result={}
        archivos = [x for x in file_path_list]
        print(file_path_list)
        print(f'Archivos a procesar: {archivos}')
        for archivo in archivos:
            try:
                print(f'Abriendo {archivo}')
                # Cargar el archivo Excel
                workbook = openpyxl.load_workbook(archivo)
                print(f'Revisando{archivo}')
                # Obtener las primeras 7 hojas
                hojas = workbook.worksheets[:7]
                fileName = archivo.split('/')[-1][:-5]
                print(f'llegando ')
                result={}
                # Iterar sobre cada hoja y realizar alguna acción 
                for hoja in hojas:
                    for i in range(3,151):
                        pieza = hoja[f'B{i}'].value
                        pieza = pieza.upper() if pieza else pieza
                        if  pieza in tipos_de_piezas:
                            if pieza  not in result.keys():
                                result[pieza]=[]
                                if pieza  not in total_result.keys():
                                    total_result[pieza]=[]
                            result[pieza].append(str(hoja[f'C{i}'].value if hoja[f'C{i}'].value else "<<VACIO>>")+  (f' [{fileName}'))
                            total_result[pieza].append(str(hoja[f'C{i}'].value if hoja[f'C{i}'].value else "<<VACIO>>")+ (f' [{fileName}]'))
                self.exportar_a_excel(nombre =f'RESUMEN_{archivo}',resultados=result,output_path=output_directory)
                
                
            except (PermissionError, FileNotFoundError, openpyxl.utils.exceptions.InvalidFileException) as e:
                print(f"Error al procesar el archivo {archivo}: {e}")
                continue
            self.exportar_a_excel(nombre =f'RESUMEN_FINAL.xlsx',resultados=total_result,output_path=output_directory)

    def exportar_a_excel(self, nombre, resultados, output_path):
        # Crear un nuevo libro de trabajo
        print(f'Exportando {nombre}...')
        workbook = openpyxl.Workbook()

        # Seleccionar la hoja activa (por defecto, la primera)
        hoja = workbook.active
        print(f'Exportando 2')
        i=1
        for key in resultados.keys():
            letra_columna = hoja.cell(row=1, column=i).column_letter
            hoja[f'{letra_columna}1']=key
            j=2
            for value in resultados[key]:
                hoja[f'{letra_columna}{j}']= str(value if value else "<<VACIO>>")
                j+=1
                if j-2 == len(resultados[key]):
                    hoja[f'{letra_columna}{j+1}'] = f'TOTAL: {j-2}'
            i+=1
        # Guardar el archivo
        print(f'Exportando {nombre}...')
        
        workbook.save(f'{output_path}/{nombre.split("/")[-1]}')